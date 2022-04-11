import openpyxl


class HandleExcel:
    def __init__(self, excel_path, sheet_name):
        """
        初始化
        :param excel_path: 所需表格路径
        :param sheet_name: 所需表格工作表
        """
        self.excel_path = excel_path
        self.sheet_name = sheet_name

    def read_excel(self):
        """
        读取表格数据
        :return: 返回表格数据
        """
        wb = openpyxl.load_workbook(self.excel_path)
        sheet = wb[self.sheet_name]
        rows = sheet.rows
        headers = [i.value for i in next(rows)]
        data = []
        for row in rows:
            cell_values = [i.value for i in row]
            dic = dict(zip(headers, cell_values))
            data.append(dic)
        return data

    def write_excel(self, row, column, value):
        """
        往表格写入数据
        :param row: 指定写入行号
        :param column: 指定写入列号
        :param value: 指定写入值
        """
        wb = openpyxl.load_workbook(self.excel_path)
        sheet = wb[self.sheet_name]
        sheet.cell(row, column, value)
        wb.save(self.excel_path)
